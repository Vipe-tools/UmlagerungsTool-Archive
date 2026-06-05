import tkinter as tk
from tkinter import messagebox, filedialog, ttk
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image, ImageTk
import win32com.client as win32
import os
import tempfile
import time
import sys
import barcode
from barcode.writer import ImageWriter
from datetime import datetime
import re

MAX_WIDTH = 270
MAX_HEIGHT = 100
PREVIEW_WIDTH = 190
PREVIEW_HEIGHT = 65

if getattr(sys, 'frozen', False):
    base_folder = sys._MEIPASS
else:
    base_folder = os.path.dirname(os.path.abspath(__file__))

excel_folder = os.path.join(base_folder, "Umlagerung")
font_file = os.path.join(base_folder, "arial.ttf")

def add_context_menu(entry):
    menu = tk.Menu(entry, tearoff=0)
    menu.add_command(label=f"{'Einfügen':25}Strg+V", command=lambda: entry.event_generate("<<Paste>>"))
    menu.add_command(label=f"{'Kopieren':25}Strg+C", command=lambda: entry.event_generate("<<Copy>>"))
    menu.add_command(label=f"{'Ausschneiden':20}Strg+X", command=lambda: entry.event_generate("<<Cut>>"))
    entry.bind("<Button-3>", lambda e: menu.tk_popup(e.x_root, e.y_root))

def refresh_file_list():
    global all_files
    if not os.path.exists(excel_folder):
        all_files = []
    else:
        all_files = [os.path.splitext(f)[0] for f in os.listdir(excel_folder) if f.lower().endswith(".xlsx")]
    excel_combo['values'] = all_files

def update_autocomplete(event=None):
    value = excel_path.get().lower()
    if value == '':
        excel_combo['values'] = all_files
    else:
        excel_combo['values'] = [f for f in all_files if value in f.lower()]

def search_excel_file(event=None):
    value = excel_path.get().strip()
    if value in all_files:
        excel_combo.set(value)
    elif value:
        filtered = [f for f in all_files if value.lower() in f.lower()]
        if filtered:
            excel_combo.set(filtered[0])
        else:
            messagebox.showerror("Fehler", "Datei nicht gefunden!")
            excel_combo.set('')
    else:
        excel_combo.set('')

def select_excel_file():
    file_path = filedialog.askopenfilename(title="Excel-Datei auswählen", filetypes=[("Excel Dateien", "*.xlsx")])
    if file_path:
        filename_no_ext = os.path.splitext(os.path.basename(file_path))[0]
        excel_path.set(filename_no_ext)
        if filename_no_ext not in all_files:
            all_files.append(filename_no_ext)
            excel_combo['values'] = all_files

def generate_barcode_from_text(event=None):
    global barcode_image, barcode_preview_tk
    fba_text = fba_entry.get().strip()
    if not fba_text:
        messagebox.showerror("Fehler", "Bitte FBA-Nummer eingeben!")
        return
    try:
        if not os.path.exists(font_file):
            messagebox.showerror("Fehler", f"Schriftdatei nicht gefunden:\n{font_file}")
            return
        writer_options = {
            "module_width": 1,
            "module_height": 40,
            "font_size": 25,
            "text_distance": 15,
            "quiet_zone": 2,
            "font_path": font_file
        }
        CODE128 = barcode.get_barcode_class('code128')
        barcode_obj = CODE128(fba_text, writer=ImageWriter())
        tmp_dir = tempfile.gettempdir()
        tmp_file_path = os.path.join(tmp_dir, "barcode.png")
        barcode_obj.write(tmp_file_path, options=writer_options)
        img = Image.open(tmp_file_path)
        ratio = min(MAX_WIDTH / img.width, MAX_HEIGHT / img.height)
        new_size = (int(img.width * ratio), int(img.height * ratio))
        img = img.resize(new_size, Image.Resampling.LANCZOS)
        barcode_image = img
        preview_ratio = min(PREVIEW_WIDTH / img.width, PREVIEW_HEIGHT / img.height)
        preview_size = (int(img.width * preview_ratio), int(img.height * preview_ratio))
        preview_img = img.resize(preview_size, Image.Resampling.LANCZOS)
        barcode_preview_tk = ImageTk.PhotoImage(preview_img)
        barcode_preview_label.configure(image=barcode_preview_tk)
        barcode_preview_label.image = barcode_preview_tk
    except Exception as e:
        messagebox.showerror("Fehler", f"Barcode konnte nicht generiert werden:\n{e}")

def process_and_print(event=None):
    global barcode_image
    try:
        number = int(number_spinbox.get())
        if not 1 <= number <= 20:
            messagebox.showerror("Fehler", "Die Zahl muss zwischen 1 und 20 liegen.")
            return
    except ValueError:
        messagebox.showerror("Fehler", "Bitte eine gültige Zahl eingeben.")
        return
    filename = excel_path.get()
    if not filename or barcode_image is None:
        messagebox.showerror("Fehler", "Bitte Excel-Datei und Barcode-Bild auswählen/generieren.")
        return
    if filename not in all_files:
        messagebox.showerror("Fehler", "Ungültige Datei ausgewählt.")
        return
    excel_file = os.path.join(excel_folder, f"{filename}.xlsx")
    if not os.path.isfile(excel_file):
        messagebox.showerror("Fehler", f"Excel-Datei nicht gefunden:\n{excel_file}")
        return
    wb = load_workbook(excel_file)
    ws = wb.active
    ws["D11"] = number
    stellplatz = stellplatz_entry.get().strip()
    if stellplatz:
        teile = re.split(r"[ \\/|+_,]+", stellplatz)
        for sp in teile:
            sp = sp.strip()
            if not sp:
                continue
            if not re.fullmatch(r"[A-Z]{2}-\d{2}", sp):
                messagebox.showerror("Fehler", f"Ungültiger Stellplatz: {sp}\nErlaubt: z.B. DD-02, AA-02")
                return
        ws["D37"] = stellplatz
    datum = datum_entry.get().strip()
    if datum:
        ws["D39"] = datum
    zusatz = zusatz_entry.get().strip()
    if zusatz:
        ws["B42"] = zusatz
    try:
        with tempfile.TemporaryDirectory() as tmpdirname:
            temp_img_path = os.path.join(tmpdirname, "barcode.png")
            temp_excel_path = os.path.join(tmpdirname, "temp.xlsx")
            barcode_image.save(temp_img_path)
            xl_img = XLImage(temp_img_path)
            xl_img.anchor = "F12"
            ws.add_image(xl_img)
            wb.save(temp_excel_path)
            excel_app = win32.DispatchEx("Excel.Application")
            excel_app.Visible = False
            excel_app.DisplayAlerts = False
            excel_app.ScreenUpdating = False
            workbook = excel_app.Workbooks.Open(temp_excel_path)
            time.sleep(0.5)
            workbook.Sheets(1).Select()
            workbook.ActiveSheet.PrintOut()
            workbook.Close(SaveChanges=False)
            excel_app.Quit()
        messagebox.showinfo("Erfolg", "Druck erfolgreich!")
    except Exception as e:
        messagebox.showerror("Fehler", f"Druck fehlgeschlagen:\n{e}")

def clear_fields():
    excel_path.set("")
    number_spinbox.delete(0, tk.END)
    number_spinbox.insert(0, "1")
    stellplatz_entry.delete(0, tk.END)
    datum_entry.delete(0, tk.END)
    datum_entry.insert(0, datetime.now().strftime("%d.%m.%Y"))
    fba_entry.delete(0, tk.END)
    barcode_preview_label.configure(image="")
    global barcode_image
    barcode_image = None
    excel_combo['values'] = all_files

def on_mousewheel_spinbox(event):
    try:
        val = int(number_spinbox.get())
    except Exception:
        val = 1
    delta = 0
    if hasattr(event, "delta"):
        if event.delta > 0:
            delta = 1
        elif event.delta < 0:
            delta = -1
    elif hasattr(event, "num"):
        if event.num == 4:
            delta = 1
        elif event.num == 5:
            delta = -1
    new = val + delta
    if new < 1:
        new = 1
    if new > 20:
        new = 20
    number_spinbox.delete(0, tk.END)
    number_spinbox.insert(0, str(new))
    return "break"

def update_date():
    current_date = datetime.now().strftime("%d.%m.%Y")
    datum_status_label.config(text=f"Aktuell: {current_date}")
    datum_entry.delete(0, tk.END)
    datum_entry.insert(0, current_date)
    root.after(3600000, update_date)

def show_info(event=None):
    filename = excel_path.get().strip()
    if not filename:
        messagebox.showerror("Fehler", "Keine Excel-Datei ausgewählt.")
        return
    excel_file = os.path.join(excel_folder, f"{filename}.xlsx")
    if not os.path.isfile(excel_file):
        messagebox.showerror("Fehler", "Excel-Datei nicht gefunden.")
        return
    try:
        wb = load_workbook(excel_file, data_only=True)
        ws = wb.active
        info_text = f"{ws['A3'].value}\n{ws['A5'].value}\n{ws['A9'].value}{ws['D9'].value}\n{ws['A13'].value}\n{ws['A15'].value}"
        messagebox.showinfo("Info zu Artikel", info_text)
    except Exception as e:
        messagebox.showerror("Fehler", f"Info konnte nicht geladen werden:\n{e}")

def show_help(event=None):
    help_text = """🟨 Anleitung – Version 1.12

Artikel suchen (SKU):
Artikelnummer eingeben → Enter oder Suchen klicken.
Das Programm zeigt automatisch alle passenden Excel-Dateien im „Umlagerung“-Ordner an.
Leeres Feld zeigt alle verfügbaren Dateien. Eingabe filtert die Vorschläge in Echtzeit.

Kartons Anzahl (1–20):
Anzahl der benötigten Kartons eingeben.
Die Zahl kann zusätzlich mit den Pfeiltasten ↑/↓ oder dem Mausrad verändert werden.

FBA Nummer:
FBA-Sendungsnummer eintragen → Enter oder Barcode generieren.
Der Barcode wird erstellt und unten als Vorschau angezeigt.

Optionale Felder:

Stellplatz: falls vorhanden, sonst leer lassen.
Mehrere Stellplätze möglich, z. B.: DD-09 / BB-34 | AA-12 + CC-03
Trennzeichen: Leerzeichen, /, \, |, +, _, ,
Jeder Stellplatz muss das Format 2 Buchstaben – 2 Zahlen haben (z. B. DD-02).

Datum: wird automatisch mit dem heutigen Datum gefüllt (kann geändert werden).

Zusatz: optionaler Kommentar zum Artikel.

Drucken:
Barcode und Angaben werden in die Excel-Datei eingefügt und automatisch über Excel gedruckt.
Nach dem Vorgang erscheint eine Meldung, ob der Druck erfolgreich war.
Neu: Tastenkürzel Strg + S für „Drucken“.

Clear:
Alle Eingaben löschen und aktuelles Datum wiederherstellen.
Die Liste der verfügbaren Excel-Dateien bleibt erhalten.

Info zu Artikel:
Zeigt Informationen zu ausgewähltem Artikel (F11 oder Info-Button).

Statusmeldungen:
F1 – Hilfe
F2 – Status Excel-Datei
F3 – Status Barcode
F11 – Info zu Artikel

🧩 Version 1.12 (16.12.2025)"""
    messagebox.showinfo("Hilfe", help_text)

def show_update_info():
    update_text = """Version 1.12 (16.12.2025) – Neues Update

Neuerungen & Verbesserungen:

Autocomplete bei Artikelsuche: Das Suchfeld zeigt jetzt automatisch alle verfügbaren Excel-Dateien im „Umlagerung“-Ordner an.

Datei-Filter: Eingabe im Suchfeld filtert die Vorschläge in Echtzeit. Leeres Feld zeigt alle Dateien.

Sicheres Auswählen: Nur vorhandene Excel-Dateien können ausgewählt werden. Ungültige Eingaben werden blockiert.

Stellplatz-Feld erweitert:

Mehrere Stellplätze möglich, z. B.: DD-02 / AA-02 | BB-03 + CC-04

Trennzeichen: Leerzeichen, /, \, |, +, -, _, ,

Validierung: Jeder Stellplatz muss 2 Buchstaben, dann -, dann 2 Zahlen haben.

Clear-Funktion verbessert: Löscht die Eingaben, behält aber die Liste der verfügbaren Excel-Dateien.

Autocomplete nach Clear: Nach Öffnen oder Clear bleibt das Suchfeld leer, die Vorschläge sind sofort verfügbar.

Allgemeine Stabilitätsverbesserungen: Fehler beim Drucken, Barcode-Generierung und Dateiauswahl wurden reduziert."""
    messagebox.showinfo("Update", update_text)

root = tk.Tk()
root.title("UmlagerungsTool v1.12")

excel_path = tk.StringVar()
barcode_image = None
barcode_preview_tk = None
all_files = []

tk.Label(root, text="Artikel Suchen (SKU):").grid(row=0, column=0, padx=5, pady=5)
excel_combo = ttk.Combobox(root, width=50, textvariable=excel_path)
excel_combo.grid(row=0, column=1, padx=5)
excel_combo.bind("<KeyRelease>", update_autocomplete)
excel_combo.bind("<Return>", search_excel_file)
tk.Button(root, text="Suchen", command=search_excel_file, bg="#90EE90").grid(row=0, column=2, padx=5)

tk.Label(root, text="Kartons Anzahl (1-20):").grid(row=1, column=0, padx=5, pady=5)
number_spinbox = tk.Spinbox(root, from_=1, to=20, width=10)
number_spinbox.grid(row=1, column=1, padx=5, sticky="w")
number_spinbox.bind("<MouseWheel>", on_mousewheel_spinbox)
number_spinbox.bind("<Button-4>", on_mousewheel_spinbox)
number_spinbox.bind("<Button-5>", on_mousewheel_spinbox)

tk.Label(root, text="FBA Nummer:").grid(row=2, column=0, padx=5, pady=5)
fba_entry = tk.Entry(root, width=30)
fba_entry.grid(row=2, column=1, padx=5, sticky="w")
fba_entry.bind("<Return>", generate_barcode_from_text)
add_context_menu(fba_entry)
tk.Button(root, text="Barcode generieren", command=generate_barcode_from_text, bg="orange").grid(row=2, column=2, padx=5)

tk.Label(root, text="Barcode-Vorschau:").grid(row=3, column=0, padx=5, pady=(0,5), sticky="e")
barcode_preview_label = tk.Label(root)
barcode_preview_label.grid(row=3, column=1, padx=5, pady=(0,5), sticky="w")

tk.Label(root, text="Optional: Stellplatz:").grid(row=4, column=0, padx=5, pady=5)
stellplatz_entry = tk.Entry(root, width=30)
stellplatz_entry.grid(row=4, column=1, padx=5, sticky="w")
add_context_menu(stellplatz_entry)
stellplatz_status_label = tk.Label(root, text="Aktuell: keine Stellplatz", fg="gray")
stellplatz_status_label.grid(row=4, column=2, padx=5, sticky="w")

tk.Label(root, text="Optional: Datum:").grid(row=5, column=0, padx=5, pady=5)
datum_entry = tk.Entry(root, width=30)
datum_entry.grid(row=5, column=1, padx=5, sticky="w")
datum_entry.insert(0, datetime.now().strftime("%d.%m.%Y"))
add_context_menu(datum_entry)
datum_status_label = tk.Label(root, text=f"Aktuell: {datetime.now().strftime('%d.%m.%Y')}", fg="gray")
datum_status_label.grid(row=5, column=2, padx=5, sticky="w")

tk.Label(root, text="Optional: Zusatz:").grid(row=6, column=0, padx=5, pady=5)
zusatz_entry = tk.Entry(root, width=30)
zusatz_entry.grid(row=6, column=1, padx=5, sticky="w")
add_context_menu(zusatz_entry)
zusatz_status_label = tk.Label(root, text="Aktuell: kein Zusatz", fg="gray")
zusatz_status_label.grid(row=6, column=2, padx=5, sticky="w")

tk.Button(root, text="Update", command=show_update_info, bg="#104E8B", fg="white").grid(row=7, column=0, pady=10, sticky="w", padx=40)
tk.Button(root, text="Drucken", command=process_and_print, bg="green", fg="white").grid(row=7, column=1, padx=5, pady=10, sticky="w")
tk.Button(root, text="Clear", command=clear_fields, bg="red", fg="white").grid(row=7, column=2, pady=10, padx=5, sticky="w")
tk.Button(root, text="Hilfe", command=show_help, bg="yellow").grid(row=7, column=2, padx=(80,5), sticky="w")
tk.Button(root, text="Info zu Artikel", command=show_info, bg="lightblue").grid(row=7, column=1, padx=(140,5), sticky="w")

root.bind_all("<Control-s>", lambda event: process_and_print())
root.bind_all("<Control-S>", lambda event: process_and_print())
root.bind_all("<F1>", show_help)
root.bind_all("<F11>", show_info)

refresh_file_list()
update_date()
root.mainloop()
