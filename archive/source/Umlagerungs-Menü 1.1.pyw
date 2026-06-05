import tkinter as tk
from tkinter import messagebox, filedialog
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import ImageGrab, Image
import win32com.client as win32
import os
import tempfile
import time
import sys

MAX_WIDTH = 270
MAX_HEIGHT = 100

# --- Ordner für Excel-Dateien ermitteln ---
if getattr(sys, 'frozen', False):
    # Für die .exe: Ordner, wo die .exe liegt
    base_folder = os.path.dirname(sys.executable)
else:
    # Für die .pyw: Ordner der Datei
    base_folder = os.path.dirname(os.path.abspath(__file__))

excel_folder = os.path.join(base_folder, "Umlagerung")

# --- Funktionen ---

def search_excel_file():
    """Excel-Datei im Ordner Umlagerung nach eingegebenem Text suchen"""
    search_text = excel_path.get().strip().lower()
    if not search_text:
        messagebox.showerror("Fehler", "Bitte Suchbegriff eingeben!")
        return

    found_file = None
    # Prüfen, ob der Ordner existiert
    if not os.path.exists(excel_folder):
        messagebox.showwarning("Warnung", f"Ordner '{excel_folder}' existiert nicht. Bitte Datei manuell auswählen.")
        select_excel_file()
        return

    for f in os.listdir(excel_folder):
        if f.lower().endswith(".xlsx") and search_text in f.lower():
            found_file = os.path.join(excel_folder, f)
            break

    if found_file:
        excel_path.set(found_file)
        messagebox.showinfo("Info", f"Datei gefunden: {os.path.basename(found_file)}")
    else:
        messagebox.showwarning("Nicht gefunden", f"Keine Datei im Ordner '{excel_folder}' gefunden, die '{search_text}' enthält.\nBitte Datei manuell auswählen.")
        select_excel_file()

def select_excel_file():
    """Dateiauswahl als Backup, wenn Suche fehlschlägt"""
    file_path = filedialog.askopenfilename(title="Excel-Datei auswählen", filetypes=[("Excel Dateien", "*.xlsx")])
    if file_path:
        excel_path.set(file_path)
    else:
        messagebox.showinfo("Abbruch", "Keine Datei ausgewählt.")

def select_barcode_image():
    """Barcode aus Zwischenablage laden und dynamisch skalieren"""
    global barcode_image
    img = ImageGrab.grabclipboard()
    if img is None:
        messagebox.showerror("Fehler", "Kein Bild in der Zwischenablage gefunden!")
        return

    ratio = min(MAX_WIDTH / img.width, MAX_HEIGHT / img.height)
    new_size = (int(img.width * ratio), int(img.height * ratio))
    img = img.resize(new_size, Image.Resampling.LANCZOS)

    barcode_image = img
    messagebox.showinfo("Info", f"Bild erfolgreich geladen ({new_size[0]}x{new_size[1]} px)!")

def process_and_print():
    """Zahl einfügen, Barcode einfügen, optionale Felder, temporäre Datei speichern und drucken"""
    global barcode_image
    try:
        number = int(number_entry.get())
        if not 1 <= number <= 20:
            messagebox.showerror("Fehler", "Die Zahl muss zwischen 1 und 20 liegen.")
            return
    except ValueError:
        messagebox.showerror("Fehler", "Bitte eine gültige Zahl eingeben.")
        return

    excel_file = excel_path.get()
    if not excel_file or barcode_image is None:
        messagebox.showerror("Fehler", "Bitte Excel-Datei und Barcode-Bild auswählen.")
        return

    wb = load_workbook(excel_file)
    ws = wb.active
    ws["D11"] = number

    # Optional: Stellplatz einfügen
    stellplatz = stellplatz_entry.get().strip()
    if stellplatz:
        ws["D37"] = stellplatz

    # Optional: Datum einfügen
    datum = datum_entry.get().strip()
    if datum:
        ws["D39"] = datum

    with tempfile.TemporaryDirectory() as tmpdirname:
        temp_img_path = os.path.join(tmpdirname, "barcode.png")
        temp_excel_path = os.path.join(tmpdirname, "temp.xlsx")
        barcode_image.save(temp_img_path)

        xl_img = XLImage(temp_img_path)
        xl_img.anchor = "F12"  # Startzelle für Barcode
        ws.add_image(xl_img)
        wb.save(temp_excel_path)

        # Excel drucken
        try:
            excel_app = win32.Dispatch("Excel.Application")
            excel_app.Visible = False
            excel_app.DisplayAlerts = False
            excel_app.ScreenUpdating = False
            workbook = excel_app.Workbooks.Open(temp_excel_path)
            time.sleep(0.5)
            workbook.Sheets(1).Select()
            workbook.ActiveSheet.PrintOut()
            workbook.Close(SaveChanges=False)
            excel_app.Quit()
            messagebox.showinfo("Fertig", "Drucken abgeschlossen!")
        except Exception as e:
            messagebox.showerror("Fehler", f"Druck fehlgeschlagen: {e}")

def clear_fields():
    """Alle Eingabefelder zurücksetzen"""
    excel_path.set("")
    number_entry.delete(0, tk.END)
    stellplatz_entry.delete(0, tk.END)
    datum_entry.delete(0, tk.END)
    global barcode_image
    barcode_image = None
    messagebox.showinfo("Info", "Alle Felder wurden geleert!")

# --- GUI ---
root = tk.Tk()
root.title("Umlagerungen zu Amazon v1.1")

excel_path = tk.StringVar()
barcode_image = None

# Suchfeld für Excel-Datei
tk.Label(root, text="Artikel Suchen (SKU):").grid(row=0, column=0, padx=5, pady=5)
tk.Entry(root, textvariable=excel_path, width=50).grid(row=0, column=1, padx=5)
tk.Button(root, text="Suchen", command=search_excel_file).grid(row=0, column=2, padx=5)

# Zahl eingeben
tk.Label(root, text="Kartons Anzahl (1-20):").grid(row=1, column=0, padx=5, pady=5)
number_entry = tk.Entry(root, width=10)
number_entry.grid(row=1, column=1, padx=5, sticky="w")

# Barcode aus Zwischenablage
tk.Label(root, text="FAB Nummer:").grid(row=2, column=0, padx=5, pady=5)
tk.Button(root, text="Bild aus Clipboard laden", command=select_barcode_image).grid(row=2, column=1, padx=5, sticky="w")

# --- Neue optionale Felder ---
# Optional: Stellplatz
tk.Label(root, text="Optional: Stellplatz:").grid(row=4, column=0, padx=5, pady=5)
stellplatz_entry = tk.Entry(root, width=30)
stellplatz_entry.grid(row=4, column=1, padx=5, sticky="w")

# Optional: Datum
tk.Label(root, text="Optional: Datum:").grid(row=5, column=0, padx=5, pady=5)
datum_entry = tk.Entry(root, width=30)
datum_entry.grid(row=5, column=1, padx=5, sticky="w")

# Drucken Button
tk.Button(root, text="Drucken", command=process_and_print, bg="green", fg="white").grid(row=6, column=1, pady=10, sticky="w")

# Clear Button
tk.Button(root, text="Clear", command=clear_fields, bg="red", fg="white").grid(row=6, column=2, pady=10, padx=5, sticky="w")

root.mainloop()
