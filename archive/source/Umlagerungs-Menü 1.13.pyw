import tkinter as tk
from tkinter import messagebox, filedialog, ttk
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image, ImageTk
import win32com.client as win32
import os
import tempfile
import time
import sys
import barcode
from barcode.writer import ImageWriter
from datetime import datetime
import re

MAX_WIDTH = 270
MAX_HEIGHT = 100
PREVIEW_WIDTH = 190
PREVIEW_HEIGHT = 65

if getattr(sys, 'frozen', False):
    base_folder = sys._MEIPASS
else:
    base_folder = os.path.dirname(os.path.abspath(__file__))

excel_folder = os.path.join(base_folder, "Umlagerung")
font_file = os.path.join(base_folder, "arial.ttf")

def add_context_menu(widget):
    menu = tk.Menu(widget, tearoff=0)
    menu.add_command(label="Einfügen\t           |  Strg+V", command=lambda: widget.event_generate("<<Paste>>"))
    menu.add_command(label="Kopieren\t           |  Strg+C", command=lambda: widget.event_generate("<<Copy>>"))
    menu.add_command(label="Ausschneiden\t  |  Strg+X", command=lambda: widget.event_generate("<<Cut>>"))
    widget.bind("<Button-3>", lambda e: menu.tk_popup(e.x_root, e.y_root))

def refresh_file_list():
    global all_files
    if not os.path.exists(excel_folder):
        all_files = []
    else:
        all_files = [os.path.splitext(f)[0] for f in os.listdir(excel_folder) if f.lower().endswith(".xlsx")]
    excel_combo['values'] = all_files

def update_autocomplete(event=None):
    value = excel_path.get().lower()
    if value == '':
        excel_combo['values'] = all_files
    else:
        excel_combo['values'] = [f for f in all_files if value in f.lower()]

def search_excel_file(event=None):
    value = excel_path.get().strip()
    if value in all_files:
        excel_combo.set(value)
    elif value:
        filtered = [f for f in all_files if value.lower() in f.lower()]
        if filtered:
            excel_combo.set(filtered[0])
        else:
            messagebox.showerror("Fehler", "Datei nicht gefunden!")
            excel_combo.set('')
    else:
        excel_combo.set('')

def autocomplete_excel():
    value = excel_path.get().strip()
    if value and value not in all_files:
        filtered = [f for f in all_files if value.lower() in f.lower()]
        if filtered:
            excel_combo.set(filtered[0])
            return filtered[0]
    return value

def select_excel_file():
    file_path = filedialog.askopenfilename(title="Excel-Datei auswählen", filetypes=[("Excel Dateien", "*.xlsx")])
    if file_path:
        filename_no_ext = os.path.splitext(os.path.basename(file_path))[0]
        excel_path.set(filename_no_ext)
        if filename_no_ext not in all_files:
            all_files.append(filename_no_ext)
            excel_combo['values'] = all_files

def generate_barcode_from_text(event=None):
    global barcode_image, barcode_preview_tk
    fba_text = fba_entry.get().strip()
    if not fba_text:
        messagebox.showerror("Fehler", "Bitte FBA-Nummer eingeben!")
        return
    try:
        if not os.path.exists(font_file):
            messagebox.showerror("Fehler", f"Schriftdatei nicht gefunden:\n{font_file}")
            return
        writer_options = {
            "module_width": 1,
            "module_height": 40,
            "font_size": 25,
            "text_distance": 15,
            "quiet_zone": 2,
            "font_path": font_file
        }
        CODE128 = barcode.get_barcode_class('code128')
        barcode_obj = CODE128(fba_text, writer=ImageWriter())
        tmp_dir = tempfile.gettempdir()
        tmp_file_path = os.path.join(tmp_dir, "barcode.png")
        barcode_obj.write(tmp_file_path, options=writer_options)
        img = Image.open(tmp_file_path)
        ratio = min(MAX_WIDTH / img.width, MAX_HEIGHT / img.height)
        img = img.resize((int(img.width * ratio), int(img.height * ratio)), Image.Resampling.LANCZOS)
        barcode_image = img
        preview_ratio = min(PREVIEW_WIDTH / img.width, PREVIEW_HEIGHT / img.height)
        preview_img = img.resize((int(img.width * preview_ratio), int(img.height * preview_ratio)), Image.Resampling.LANCZOS)
        barcode_preview_tk = ImageTk.PhotoImage(preview_img)
        barcode_preview_label.configure(image=barcode_preview_tk)
        barcode_preview_label.image = barcode_preview_tk
    except Exception as e:
        messagebox.showerror("Fehler", f"Barcode konnte nicht generiert werden:\n{e}")

def process_and_print(event=None):
    global barcode_image
    filename = autocomplete_excel()
    excel_path.set(filename)
    if fba_entry.get().strip() and barcode_image is None:
        generate_barcode_from_text()
    try:
        number = int(number_spinbox.get())
        if not 1 <= number <= 20:
            messagebox.showerror("Fehler", "Die Zahl muss zwischen 1 und 20 liegen.")
            return
    except ValueError:
        messagebox.showerror("Fehler", "Bitte eine gültige Zahl eingeben.")
        return
    if not filename or barcode_image is None:
        messagebox.showerror("Fehler", "Bitte Excel-Datei und Barcode-Bild auswählen/generieren.")
        return
    if filename not in all_files:
        messagebox.showerror("Fehler", "Ungültige Datei ausgewählt.")
        return
    excel_file = os.path.join(excel_folder, f"{filename}.xlsx")
    if not os.path.isfile(excel_file):
        messagebox.showerror("Fehler", f"Excel-Datei nicht gefunden:\n{excel_file}")
        return
    wb = load_workbook(excel_file)
    ws = wb.active
    ws["D11"] = number
    stellplatz = stellplatz_entry.get().strip()
    if stellplatz:
        teile = re.split(r"[ \\/|+_,]+", stellplatz)
        for sp in teile:
            sp = sp.strip()
            if sp and not re.fullmatch(r"[A-Z]{2}-\d{2}", sp):
                messagebox.showerror("Fehler", f"Ungültiger Stellplatz: {sp}")
                return
        ws["D37"] = stellplatz
    datum = datum_entry.get().strip()
    if datum:
        ws["D39"] = datum
    zusatz = zusatz_entry.get().strip()
    if zusatz:
        ws["B42"] = zusatz
        zusatz_status_label.config(text=f"Aktuell: {zusatz}")
    try:
        with tempfile.TemporaryDirectory() as tmpdirname:
            temp_img_path = os.path.join(tmpdirname, "barcode.png")
            temp_excel_path = os.path.join(tmpdirname, "temp.xlsx")
            barcode_image.save(temp_img_path)
            xl_img = XLImage(temp_img_path)
            xl_img.anchor = "F12"
            ws.add_image(xl_img)
            wb.save(temp_excel_path)
            excel_app = win32.DispatchEx("Excel.Application")
            excel_app.Visible = False
            excel_app.DisplayAlerts = False
            excel_app.ScreenUpdating = False
            workbook = excel_app.Workbooks.Open(temp_excel_path)
            time.sleep(0.5)
            workbook.ActiveSheet.PrintOut()
            workbook.Close(SaveChanges=False)
            excel_app.Quit()
        messagebox.showinfo("Erfolg", "Druck erfolgreich!")
    except Exception as e:
        messagebox.showerror("Fehler", f"Druck fehlgeschlagen:\n{e}")

def clear_fields():
    excel_path.set("")
    number_spinbox.delete(0, tk.END)
    number_spinbox.insert(0, "1")
    stellplatz_entry.delete(0, tk.END)
    datum_entry.delete(0, tk.END)
    datum_entry.insert(0, datetime.now().strftime("%d.%m.%Y"))
    fba_entry.delete(0, tk.END)
    zusatz_entry.delete(0, tk.END)
    zusatz_status_label.config(text="Aktuell: kein Zusatz", fg="gray")
    barcode_preview_label.configure(image="")
    global barcode_image
    barcode_image = None
    excel_combo['values'] = all_files

def on_mousewheel_spinbox(event):
    try:
        val = int(number_spinbox.get())
    except Exception:
        val = 1
    delta = 1 if getattr(event, "delta", 0) > 0 or getattr(event, "num", None) == 4 else -1
    new = max(1, min(20, val + delta))
    number_spinbox.delete(0, tk.END)
    number_spinbox.insert(0, str(new))
    return "break"

def update_date():
    current_date = datetime.now().strftime("%d.%m.%Y")
    datum_status_label.config(text=f"Aktuell: {current_date}")
    datum_entry.delete(0, tk.END)
    datum_entry.insert(0, current_date)
    root.after(3600000, update_date)

def show_long_text(title, text):
    window = tk.Toplevel(root)
    window.title(title)
    window.geometry("600x400")
    window.grab_set()
    
    scrollbar = tk.Scrollbar(window)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
    
    text_widget = tk.Text(window, wrap=tk.WORD, yscrollcommand=scrollbar.set)
    text_widget.pack(expand=True, fill=tk.BOTH)
    text_widget.insert(tk.END, text)
    text_widget.config(state=tk.DISABLED)
    
    scrollbar.config(command=text_widget.yview)

def show_help(event=None):
    long_help_text = """🟨 Anleitung – Version 1.13

Artikel suchen (SKU):
Artikelnummer eingeben → Enter oder Suchen klicken.
Das Programm zeigt automatisch alle passenden Excel-Dateien im „Umlagerung“-Ordner an.
Leeres Feld zeigt alle verfügbaren Dateien. Eingabe filtert die Vorschläge in Echtzeit, auch bei Sonderzeichen.

Kartons Anzahl (1–20):
Anzahl der benötigten Kartons eingeben.
Die Zahl kann zusätzlich mit den Pfeiltasten ↑/↓ oder dem Mausrad verändert werden.

FBA Nummer:
FBA-Sendungsnummer eintragen → Enter oder Barcode generieren.
Der Barcode wird erstellt und unten als Vorschau angezeigt. Anzeigeprobleme bei großen Barcodes wurden behoben.

Optionale Felder:

Stellplatz: falls vorhanden, sonst leer lassen.
Mehrere Stellplätze möglich, z. B.: DD-09 / BB-34 | AA-12 + CC-03
Trennzeichen: Leerzeichen, /, , |, +, _, ,
Jeder Stellplatz muss das Format 2 Buchstaben – 2 Zahlen haben (z. B. DD-02).

Datum: wird automatisch mit dem heutigen Datum gefüllt (kann geändert werden). Korrekte Aktualisierung nach Clear oder manuellem Eintrag wurde sichergestellt.

Zusatz: optionaler Kommentar zum Artikel. Nach dem Drucken wird der aktuelle Zusatz im Status angezeigt.

Drucken:
Barcode und Angaben werden in die Excel-Datei eingefügt und automatisch über Excel gedruckt.
Nach dem Vorgang erscheint eine Meldung, ob der Druck erfolgreich war.
Neu: Tastenkürzel Strg + S für „Drucken“.

Clear:
Alle Eingaben löschen und aktuelles Datum wiederherstellen.
Die Liste der verfügbaren Excel-Dateien bleibt erhalten. Alle Vorschläge im Suchfeld werden sofort angezeigt.

Info zu Artikel:
Zeigt Informationen zu ausgewähltem Artikel (F11 oder Info-Button). Autovervollständigung bei der Artikelsuche wird ebenfalls berücksichtigt.

Statusmeldungen:
F1 – Hilfe
F11 – Info zu Artikel

🧩 Version 1.13 (29.04.2026)
"""
    show_long_text("Hilfe", long_help_text)

def show_update_info():
    long_update_text = """Version 1.13 (29.04.2026) – Installer & Update verbessert

Neuerungen & Verbesserungen:

Neue Update-Funktion:
Das Programm kann jetzt direkt über den Installer aktualisiert werden, ohne vorherige Deinstallation.

Saubere Installation:
Alte Programmdateien werden beim Update automatisch entfernt und durch neue ersetzt.

Keine Geisterdateien mehr:
Veraltete Dateien früherer Versionen bleiben nicht mehr im Installationsordner zurück.

Bestehende Installation wird erkannt:
Wenn das Tool bereits installiert ist, wird automatisch das vorhandene Verzeichnis genutzt.

Desktop- und Startmenü-Verknüpfungen bleiben erhalten:
Beim Update müssen diese nicht neu erstellt werden.

Stabilerer Installer:
Installationsprobleme auf einigen PCs wurden reduziert.

Allgemeine Verbesserungen:
Optimierungen bei Update-Ablauf, Dateiersetzung und Programmstart.
"""
    messagebox.showinfo("Update", long_update_text)

# Artikel info
def show_info(event=None):
    filename = autocomplete_excel()
    excel_path.set(filename)

    if not filename:
        messagebox.showerror("Fehler", "Keine Excel-Datei ausgewählt.")
        return

    excel_file = os.path.join(excel_folder, f"{filename}.xlsx")

    if not os.path.isfile(excel_file):
        messagebox.showerror("Fehler", "Excel-Datei nicht gefunden.")
        return

    wb = load_workbook(excel_file, data_only=True)
    ws = wb.active

    info_text = f"{ws['A3'].value}\n{ws['A5'].value}\n{ws['A9'].value}{ws['D9'].value}\n{ws['A13'].value}\n{ws['A15'].value}"

    messagebox.showinfo("Info zu Artikel", info_text)

root = tk.Tk()
root.title("UmlagerungsTool v1.13")

excel_path = tk.StringVar()
barcode_image = None
barcode_preview_tk = None
all_files = []

tk.Label(root, text="Artikel Suchen (SKU):").grid(row=0, column=0, padx=5, pady=5)
excel_combo = ttk.Combobox(root, width=50, textvariable=excel_path)
excel_combo.grid(row=0, column=1, padx=5)
excel_combo.bind("<KeyRelease>", update_autocomplete)
excel_combo.bind("<Return>", search_excel_file)
add_context_menu(excel_combo)
tk.Button(root, text="Suchen", command=search_excel_file, bg="#90EE90").grid(row=0, column=2, padx=5)

tk.Label(root, text="Kartons Anzahl (1-20):").grid(row=1, column=0, padx=5, pady=5)
number_spinbox = tk.Spinbox(root, from_=1, to=20, width=10)
number_spinbox.grid(row=1, column=1, padx=5, sticky="w")
number_spinbox.bind("<MouseWheel>", on_mousewheel_spinbox)
number_spinbox.bind("<Button-4>", on_mousewheel_spinbox)
number_spinbox.bind("<Button-5>", on_mousewheel_spinbox)

tk.Label(root, text="FBA Nummer:").grid(row=2, column=0, padx=5, pady=5)
fba_entry = tk.Entry(root, width=30)
fba_entry.grid(row=2, column=1, padx=5, sticky="w")
fba_entry.bind("<Return>", generate_barcode_from_text)
add_context_menu(fba_entry)
tk.Button(root, text="Barcode generieren", command=generate_barcode_from_text, bg="orange").grid(row=2, column=2, padx=5)

tk.Label(root, text="Barcode-Vorschau:").grid(row=3, column=0, padx=5, pady=(0,5), sticky="e")
barcode_preview_label = tk.Label(root)
barcode_preview_label.grid(row=3, column=1, padx=5, pady=(0,5), sticky="w")

tk.Label(root, text="Optional: Stellplatz:").grid(row=4, column=0, padx=5, pady=5)
stellplatz_entry = tk.Entry(root, width=30)
stellplatz_entry.grid(row=4, column=1, padx=5, sticky="w")
add_context_menu(stellplatz_entry)
stellplatz_status_label = tk.Label(root, text="Aktuell: keine Stellplatz", fg="gray")
stellplatz_status_label.grid(row=4, column=2, padx=5, sticky="w")

tk.Label(root, text="Optional: Datum:").grid(row=5, column=0, padx=5, pady=5)
datum_entry = tk.Entry(root, width=30)
datum_entry.grid(row=5, column=1, padx=5, sticky="w")
datum_entry.insert(0, datetime.now().strftime("%d.%m.%Y"))
add_context_menu(datum_entry)
datum_status_label = tk.Label(root, text=f"Aktuell: {datetime.now().strftime('%d.%m.%Y')}", fg="gray")
datum_status_label.grid(row=5, column=2, padx=5, sticky="w")

tk.Label(root, text="Optional: Zusatz:").grid(row=6, column=0, padx=5, pady=5)
zusatz_entry = tk.Entry(root, width=30)
zusatz_entry.grid(row=6, column=1, padx=5, sticky="w")
add_context_menu(zusatz_entry)
zusatz_status_label = tk.Label(root, text="Aktuell: kein Zusatz", fg="gray")
zusatz_status_label.grid(row=6, column=2, padx=5, sticky="w")

tk.Button(root, text="Update", command=show_update_info, bg="#104E8B", fg="white").grid(row=7, column=0, pady=10, sticky="w", padx=40)
tk.Button(root, text="Drucken", command=process_and_print, bg="green", fg="white").grid(row=7, column=1, padx=5, pady=10, sticky="w")
tk.Button(root, text="Clear", command=clear_fields, bg="red", fg="white").grid(row=7, column=2, pady=10, padx=5, sticky="w")
tk.Button(root, text="Hilfe", command=show_help, bg="yellow").grid(row=7, column=2, padx=(80,5), sticky="w")
tk.Button(root, text="Info zu Artikel", command=show_info, bg="lightblue").grid(row=7, column=1, padx=(140,5), sticky="w")

root.bind_all("<Control-s>", lambda event: process_and_print())
root.bind_all("<Control-S>", lambda event: process_and_print())
root.bind_all("<F1>", show_help)
root.bind_all("<F11>", show_info)

refresh_file_list()
update_date()
root.mainloop()
