import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import os
import sys
import shutil
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from datetime import datetime
import re
import traceback
import barcode
from barcode.writer import ImageWriter

# =========================
# SAFE PATH
# =========================
try:
    if getattr(sys, 'frozen', False):
        base_path = os.path.dirname(sys.executable)
    else:
        base_path = os.path.dirname(os.path.abspath(__file__))

    BASE_DIR = os.path.join(base_path, "Umlagerung")
    os.makedirs(BASE_DIR, exist_ok=True)

except Exception as e:
    print("PATH ERROR:", e)
    input()
    sys.exit()

TEMPLATE_NAME = "%Vorlage%.xlsx"
TEMPLATE_PATH = os.path.join(BASE_DIR, TEMPLATE_NAME)

all_files = []
selected_file = None
search_after_id = None

FIELDS = [
    "Artikelname",
    "SKU",
    "Amazon SKU",
    "Anzahl im Karton",
    "Kartongröße",
    "Kartongewicht"
]

def sanitize_filename(text):
    return re.sub(r'[\\/*?:"<>|]', "_", str(text))


def transform_barcode_text(text):
    if not text:
        return text

    mapping = {
        "Z": "Y",
        "z": "y",
        "y": "z",
        "Y": "Z",
        "-": "/"
    }

    return "".join(mapping.get(c, c) for c in str(text))

CHECK_MAP = {
    "Etikettieren": ("C19", "E19"),
    "Siegel": ("C21", "E21"),
    "Aufkleber SP": ("C23", "E23"),
    "Umpacken": ("C25", "E25"),
    "Bänder ab": ("C27", "E27"),
    "Gefahrgut": ("C29", "E29"),
}

# =========================
# CLIPBOARD / RIGHT CLICK
# =========================
def bind_clipboard(widget):
    widget.bind("<Control-a>", lambda e: (widget.select_range(0, tk.END), "break"))
    widget.bind("<Control-c>", lambda e: (root.clipboard_clear(), root.clipboard_append(widget.selection_get()), "break"))
    widget.bind("<Control-v>", lambda e: (widget.insert(tk.INSERT, root.clipboard_get()), "break"))

def create_context_menu(widget):
    menu = tk.Menu(widget, tearoff=0)
    menu.add_command(label="Kopieren", command=lambda: widget.event_generate("<<Copy>>"))
    menu.add_command(label="Einfügen", command=lambda: widget.event_generate("<<Paste>>"))
    menu.add_command(label="Alles markieren", command=lambda: widget.select_range(0, tk.END))

    def show(event):
        menu.tk_popup(event.x_root, event.y_root)

    widget.bind("<Button-3>", show)

def apply_input_features(widget):
    if isinstance(widget, tk.Entry):
        bind_clipboard(widget)
        create_context_menu(widget)

def add_global_clipboard_support(window):
    for child in window.winfo_children():
        try:
            apply_input_features(child)
        except:
            pass

# =========================
# FILE LOAD
# =========================
def load_files():
    global all_files
    try:
        all_files = [
            f for f in os.listdir(BASE_DIR)
            if f.endswith(".xlsx") and "%Vorlage%" not in f
        ]
        search_combo["values"] = all_files
    except Exception as e:
        messagebox.showerror("Fehler", f"Load Files Error:\n{e}")

# =========================
# SEARCH
# =========================
def refresh_search(event=None):
    global search_after_id

    if search_after_id:
        root.after_cancel(search_after_id)

    search_after_id = root.after(150, apply_search)

def apply_search():
    q = search_var.get().lower().strip()

    if not q:
        search_combo["values"] = all_files
        return

    search_combo["values"] = [f for f in all_files if q in f.lower()]

def select_best(event=None):
    global selected_file

    val = search_combo.get().strip().lower()
    matches = [f for f in all_files if val in f.lower()]

    if matches:
        selected_file = matches[0]
        status_label.config(text=f"Ausgewählt: {selected_file}")

def select_dropdown(event=None):
    global selected_file

    val = search_combo.get()

    if val in all_files:
        selected_file = val
        status_label.config(text=f"Ausgewählt: {selected_file}")
    else:
        selected_file = None
        status_label.config(text="Ungültige Auswahl")

# =========================
# INFO
# =========================
def safe(v):
    return "" if v is None else v

def show_info():
    if not selected_file:
        messagebox.showerror("Fehler", "Kein Artikel gewählt")
        return

    try:
        path = os.path.join(BASE_DIR, selected_file)
        wb = load_workbook(path, data_only=True)
        ws = wb.active

        text = f"""
Artikel: {safe(ws['A3'].value)}
SKU: {safe(ws['A5'].value)}
Amazon SKU: {safe(ws['A7'].value)}
Anzahl: {safe(ws['D9'].value)}
Größe: {safe(ws['A13'].value)}
Gewicht: {safe(ws['A15'].value)}
"""
        messagebox.showinfo("Artikel Info", text)

    except Exception as e:
        messagebox.showerror("Fehler", str(e))

# =========================
# NEW ARTICLE
# =========================
def new_article():
    win = tk.Toplevel(root)
    win.title("Neuer Artikel")

    entries = {}
    checks = {}

    r = 0

    for f in FIELDS:

        label_text = f

        if f == "Kartongröße":
            label_text = "Kartongröße (cm)"

        elif f == "Kartongewicht":
            label_text = "Kartongewicht (kg)"

        tk.Label(
            win,
            text=label_text,
            font=("Arial", 10, "bold")
        ).grid(
            row=r,
            column=0,
            sticky="w",
            padx=10,
            pady=4
        )

        e = tk.Entry(win, width=40)

        e.grid(
            row=r,
            column=1,
            padx=10,
            pady=4
        )

        entries[f] = e
        r += 1
    

    for name in CHECK_MAP:
        var = tk.IntVar()
        tk.Checkbutton(win, text=name, variable=var).grid(row=r, column=0, sticky="w")
        checks[name] = var
        r += 1

    add_global_clipboard_support(win)

    def save():
        try:
            title = entries["Artikelname"].get().strip()
            sku_raw = entries["SKU"].get().strip()

            sku_barcode = transform_barcode_text(sku_raw)
            sku_barcode_file = sanitize_filename(sku_barcode)
            sku_filename = sanitize_filename(sku_raw)

            amazon = entries["Amazon SKU"].get().strip()
            try:
                font_file = os.path.join(base_path, "arial.ttf")

                if os.path.exists(font_file):

                    writer_options = {
                        "module_width": 1,
                        "module_height": 40,
                        "font_size": 25,
                        "text_distance": 15,
                        "quiet_zone": 2,
                        "font_path": font_file
                    }

                    CODE128 = barcode.get_barcode_class("code128")

                    barcode_obj = CODE128(
                        sku_barcode,
                        writer=ImageWriter()
                    )

                    barcode_file = os.path.join(
                        BASE_DIR,
                        f"{sku_barcode_file}_barcode.png"
                    )

                    with open(barcode_file, "wb") as fp:
                        barcode_obj.write(
                            fp,
                            options=writer_options
                        )

            except Exception as e:
                messagebox.showerror(
                    "Barcode Fehler",
                    str(e)
                )

            if not title or not sku_barcode:
                messagebox.showerror("Fehler", "Artikelname + SKU fehlt")
                return

            safe_title = re.sub(r'[\\/*?:"<>|]', "", title)

            filename = " --- ".join([
                sku_filename,
                sanitize_filename(amazon),
                sanitize_filename(safe_title)
            ]) + ".xlsx"
            path = os.path.join(BASE_DIR, filename)

            if os.path.exists(TEMPLATE_PATH):
                shutil.copyfile(TEMPLATE_PATH, path)
                wb = load_workbook(path)
            else:
                from openpyxl import Workbook
                wb = Workbook()

            ws = wb.active

            ws["A3"] = f"Artikel: {title}"
            ws["A5"] = f"SKU: {sku_raw}"
            ws["A7"] = f"Amazon SKU: {amazon}"

            ws["D9"] = entries["Anzahl im Karton"].get()
            ws["A13"] = f"Kartongröße: {entries['Kartongröße'].get().strip()} cm"
            ws["A15"] = f"Kartongewicht: {entries['Kartongewicht'].get().strip()} kg"

            CODE128 = barcode.get_barcode_class("code128")
            barcode_obj = CODE128(sku_barcode, writer=ImageWriter())

            barcode_base = os.path.join(BASE_DIR, f"{sku_barcode_file}_barcode")

            barcode_file = barcode_obj.save(barcode_base)

            barcode_img = XLImage(barcode_file)
            barcode_img.width = 220
            barcode_img.height = 80
            barcode_img.anchor = "D5"

            ws.add_image(barcode_img)

            for name, (c, e) in CHECK_MAP.items():
                if checks[name].get():
                    ws[c] = "X"
                    ws[e] = ""
                else:
                    ws[c] = ""
                    ws[e] = "X"

            ws["A39"] = "Erstellungsdatum:"
            ws["D39"] = datetime.now().strftime("%d.%m.%Y")

            wb.save(path)
            load_files()
            win.destroy()

        except Exception as e:
            messagebox.showerror("Fehler", str(e))

    tk.Button(win, text="Speichern", bg="green", fg="white", width=20, command=save).grid(row=r, column=0)
    tk.Button(win, text="Abbrechen", bg="red", fg="white", width=20, command=win.destroy).grid(row=r, column=1)

# =========================
# DELETE
# =========================
def delete_article():
    global selected_file

    if not selected_file:
        messagebox.showerror("Fehler", "Kein Artikel gewählt")
        return

    path = os.path.join(BASE_DIR, selected_file)

    if messagebox.askyesno("Löschen", f"Wirklich löschen?\n\n{selected_file}"):
        os.remove(path)
        selected_file = None
        load_files()
        search_var.set("")
        search_combo.set("")
        status_label.config(text="Kein Artikel gewählt")

# =========================
# EDIT
# =========================
def edit_article():
    if not selected_file:
        messagebox.showerror("Fehler", "Kein Artikel gewählt")
        return

    try:
        path = os.path.join(BASE_DIR, selected_file)
        wb = load_workbook(path)
        ws = wb.active

        win = tk.Toplevel(root)
        win.title("Artikel bearbeiten")

        def clean(value, prefix):
            if not value:
                return ""

            text = str(value).replace(prefix, "").strip()

            if prefix == "Kartongröße:":
                text = re.sub(r"\s*cm\s*$", "", text, flags=re.IGNORECASE)

            if prefix == "Kartongewicht:":
                text = re.sub(r"\s*kg\s*$", "", text, flags=re.IGNORECASE)

            return text.strip()

        entries = {}
        checks = {}

        data = {
            "Artikelname": clean(ws["A3"].value, "Artikel:"),
            "SKU": clean(ws["A5"].value, "SKU:"),
            "Amazon SKU": clean(ws["A7"].value, "Amazon SKU:"),
            "Anzahl im Karton": ws["D9"].value,
            "Kartongröße": clean(ws["A13"].value, "Kartongröße:"),
            "Kartongewicht": clean(ws["A15"].value, "Kartongewicht:"),
        }

        r = 0

        for k, v in data.items():

            label_text = k

            if k == "Kartongröße":
                label_text = "Kartongröße (cm)"

            elif k == "Kartongewicht":
                label_text = "Kartongewicht (kg)"

            tk.Label(
                win,
                text=label_text,
                font=("Arial", 10, "bold")
            ).grid(
                row=r,
                column=0,
                sticky="w",
                padx=10,
                pady=4
            )

            e = tk.Entry(win, width=40)

            e.insert(
                0,
                "" if v is None else str(v)
            )

            e.grid(
                row=r,
                column=1,
                padx=10,
                pady=4
            )

            entries[k] = e
            r += 1

        for name, (c, e) in CHECK_MAP.items():
            var = tk.IntVar()
            var.set(1 if ws[c].value == "X" else 0)
            tk.Checkbutton(win, text=name, variable=var).grid(row=r, column=0, columnspan=2, sticky="w", padx=20, pady=2)
            checks[name] = var
            r += 1

        add_global_clipboard_support(win)

        def save():
            try:
                ws["A3"] = f"Artikel: {entries['Artikelname'].get().strip()}"
                ws["A5"] = f"SKU: {entries['SKU'].get().strip()}"
                ws["A7"] = f"Amazon SKU: {entries['Amazon SKU'].get().strip()}"

                ws["D9"] = entries["Anzahl im Karton"].get().strip()
                ws["A13"] = f"Kartongröße: {entries['Kartongröße'].get().strip()} cm"
                ws["A15"] = f"Kartongewicht: {entries['Kartongewicht'].get().strip()} kg"

                for name, (c, e) in CHECK_MAP.items():
                    if checks[name].get():
                        ws[c] = "X"
                        ws[e] = ""
                    else:
                        ws[c] = ""
                        ws[e] = "X"

                wb.save(path)
                load_files()
                win.destroy()

            except Exception as e:
                messagebox.showerror("Fehler", str(e))

        tk.Button(
            win,
            text="Speichern",
            bg="green",
            fg="white",
            width=20,
            command=save
        ).grid(row=r, column=0, padx=10, pady=10)

        tk.Button(
            win,
            text="Abbrechen",
            bg="red",
            fg="white",
            width=20,
            command=win.destroy
        ).grid(row=r, column=1, padx=10, pady=10)

    except Exception as e:
        messagebox.showerror("Start Fehler", str(e))

def show_help():

    messagebox.showinfo(
        "Help",
        """Artikel Verwaltung

Artikel Info
Zeigt die Daten des gewählten Artikels.

Bearbeiten
Bearbeitet den gewählten Artikel.

Neuer Artikel erstellen
Legt einen neuen Artikel an.

Artikel Löschen
Löscht den gewählten Artikel.

Export ZIP
Exportiert alle Artikel.

Import ZIP
Importiert eine ZIP-Datei.

STRG+A = Alles markieren
STRG+C = Kopieren
STRG+V = Einfügen
Rechtsklick = Menü"""
    )

def on_close():
    try:
        for f in os.listdir(BASE_DIR):
            if f.endswith("_barcode.png"):
                try:
                    os.remove(os.path.join(BASE_DIR, f))
                except:
                    pass
    finally:
        root.destroy()

# =========================
# UI
# =========================
root = tk.Tk()
root.protocol("WM_DELETE_WINDOW", on_close)
root.title("Artikel Verwaltung")
root.geometry("720x520")

search_var = tk.StringVar()

search_combo = ttk.Combobox(root, textvariable=search_var, width=70)
search_combo.pack(pady=10)

search_combo.bind("<KeyRelease>", refresh_search)
search_combo.bind("<Return>", select_best)
search_combo.bind("<<ComboboxSelected>>", select_dropdown)

status_label = tk.Label(root, text="Kein Artikel gewählt")
status_label.pack()

tk.Button(root, text="Artikel Info", bg="lightblue", width=25, command=show_info).pack(pady=3)
tk.Button(root, text="Bearbeiten", bg="darkblue", fg="white", width=25, command=edit_article).pack(pady=3)
tk.Button(root, text="Neuer Artikel erstellen", bg="green", fg="white", width=25, command=new_article).pack(pady=3)
tk.Button(root, text="Artikel Löschen", bg="red", fg="white", width=25, command=delete_article).pack(pady=3)

tk.Button(root, text="Export ZIP", width=25, command=lambda: shutil.make_archive("export", "zip", BASE_DIR)).pack(pady=3)
tk.Button(root, text="Import ZIP", width=25, command=lambda: shutil.unpack_archive(filedialog.askopenfilename(), BASE_DIR)).pack(pady=3)
tk.Button(root, text="Help", bg="yellow", width=25, command=show_help).pack(pady=3)
tk.Button(root, text="Zurück", bg="gray", fg="white", width=25, command=on_close).pack(pady=10)

try:
    load_files()
    root.mainloop()
except Exception:
    traceback.print_exc()
    input("Fehler - Enter")